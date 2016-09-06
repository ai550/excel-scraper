from openpyxl import load_workbook
import os

# Set things up, prepare workbook and sheet
mrf_log = load_workbook(filename = 'C:\\Users\\nk91009578.NCOC\\Desktop\\MRFs\\2013\\NCPOC MRF 0428.xlsx')
sheet = mrf_log.active
datalist = [] # for each workbook, we will have datalist - a formatted data


columns = {1:"A", 2:"B", 3:"C", 4:"D", 5:"E", 6:"F", 7:"G", 8:"H", 9:"I", 10:"J"}
metadata = {}

# check if all fields are in place for header
def get_metadata(sheet):
	
	initial_row_1 = 3
	initial_row_2 = 4
	initial_col_1 = 5
	initial_col_2 = 6

	from_loc = sheet[columns[initial_col_1] + str(initial_row_1)].value
	to_loc = sheet[columns[initial_col_2] + str(initial_row_1)].value
	mrf_n = sheet[columns[initial_col_1] + str(initial_row_2)].value
	date_n = sheet[columns[initial_col_2] + str(initial_row_2)].value

	if(from_loc == None or mrf_n == None):
		# step left
		initial_col_1 -= 1
		from_loc = sheet[columns[initial_col_1] + str(initial_row_1)].value
		mrf_n = sheet[columns[initial_col_1] + str(initial_row_2)].value
		# check again
		if(from_loc == None or mrf_n == None):
			# step right
			initial_col_1 += 2
			from_loc = sheet[columns[initial_col_1] + str(initial_row_1)].value
			mrf_n = sheet[columns[initial_col_1] + str(initial_row_2)].value

	if(to_loc == None or date_n == None):
		# step left
		initial_col_2 -= 1
		to_loc = sheet[columns[initial_col_2] + str(initial_row_1)].value
		date_n = sheet[columns[initial_col_2] + str(initial_row_2)].value
		# check again
		if(to_loc == None or date_n == None):
			# step right
			initial_col_2 += 2
			to_loc = sheet[columns[initial_col_2] + str(initial_row_1)].value
			date_n = sheet[columns[initial_col_2] + str(initial_row_2)].value
	date_n = date_n.strftime('%d/%m/%Y')
	metadata['fl'] = from_loc
	metadata['tl'] = to_loc
	metadata['mrf'] = mrf_n
	metadata['date'] = date_n

def scan_excel(search_word, sheet, filename):
	for row in range(sheet.max_row):
		for col in columns:
			cell_name = "{}{}".format(col, row + 1)
			cell_value = sheet[cell_name].value
			if(search_word == cell_value):
				print(filename)

def scan_dir(directory, search_word):
	for file in os.listdir(directory):
	    if(os.path.isfile(directory + "\\" + file)):
	        if file.endswith(".xlsx"):
	        	file_name = directory + "\\" + file
	        	workbook = load_workbook(filename = file_name)
	        	sheet = workbook.active
	        	scan_excel(search_word, sheet, file_name)
	    if(os.path.isdir(directory + "\\" + file)):
	        print("Directory: " + file)
	        scan_dir(directory + "\\" + file, search_word)

def get_consumables(sheet, col, row):
	row += 2
	description = sheet[columns[col] + str(row)].value
	quantity = sheet[columns[10] + str(row)].value		
	while(description != None):
		n_quantity = ""
		for i in list(quantity):
			if(i.isnumeric()):
				n_quantity = 
		datalist.append(['Consumables', description, '', '', quantity, metadata['tl'], metadata['fl'], metadata['date'], metadata['mrf']])
		# print('Consumables', description, '', '', quantity, metadata['tl'], metadata['fl'], metadata['date'], metadata['mrf'])
		row += 1
		description = sheet[columns[col] + str(row)].value
		quantity = sheet[columns[10] + str(row)].value

def stumble_upon(sheet, something):
	result = ""
	outer_break = False
	for row in range(sheet.max_row):
		for col in columns:
			cell_name = "{}{}".format(columns[col], row + 1)
			if(str(sheet[cell_name].value) == something or sheet[cell_name].value):
				# print("Stumbled upon {} at {}".format(something, cell_name))
				result = [col, row + 1]
				outer_break = True
				break
		if(outer_break):
			break
	return result

def get_assets_and_commodities(sheet, s_col, d_col, row):
	row += 1
	d_cell_value = sheet[columns[s_col + 1] + str(row)].value
	s_cell_value = sheet[columns[s_col] + str(row)].value
	# d_cell_value = sheet[columns[d_col] + str(row)].value
	while(s_cell_value != None or d_cell_value != None):
		datalist.append(['Asset', d_cell_value, '', s_cell_value, 1, metadata['tl'], metadata['fl'], metadata['date'], metadata['mrf']])
		# print('Asset', d_cell_value, '', s_cell_value, 1, metadata['tl'], metadata['fl'], metadata['date'], metadata['mrf'])
		row += 1
		s_cell_value = sheet[columns[s_col] + str(row)].value
		d_cell_value = sheet[columns[s_col + 1] + str(row)].value

# For test purposes
# stumble_upon(sheet, 'Assets & Commodities')
# stumble_upon(sheet, 'Serial Number')
# stumble_upon(sheet, 'Description')
# stumble_upon(sheet, 'Status')

get_metadata(sheet)

sn = stumble_upon(sheet, 'Serial Number')
ds = stumble_upon(sheet, 'Description')

get_assets_and_commodities(sheet, sn[0], ds[0], sn[1])

r = stumble_upon(sheet, 'Consumables')
col = r[0]
row = r[1]
get_consumables(sheet, col, row)

def lead_pipe(data, sheet):
	# first, let's check the last used row a.k.a max row
	print('Last row: ' + str(sheet.max_row) + ' --- Next row: ' + str(sheet.max_row + 1))
	print('Number of columns used: ' + str(sheet.max_column))

# mrf_logger = load_workbook(filename = 'C:\\Users\\nk91009578.NCOC\\Desktop\\MRF_logger - Copy.xlsx')
# mrf_sheet = mrf_logger.active

# lead_pipe('data', mrf_sheet)

print(datalist)