from openpyxl import load_workbook


def getPath():
    path = input("--> Please enter path to your file\n")
    print("--> Fetching data from", path)
    return path

def normalizeData(wb):
    sheet = wb.active
    cell = input("--> Please enter clugged cell name\n")
    data = str(sheet.cell(cell).value)
    data_split = data.split(",")
    for d in data_split:
        print(d.strip())

path = getPath()
wb = load_workbook(filename=path)

normalizeData(wb)
    
