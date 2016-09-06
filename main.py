from openpyxl import load_workbook
from datetime import date, datetime
from metadata import *
from assets_retriever import * 
import os

mrfs = []


def dir_search(directory):
    for file in os.listdir(directory):
        if(os.path.isfile(directory + '\\' + file)):
            if file.endswith('.xlsx'):
            	# print('--- Excel File: ' + file)
            	# print('------ Path: ' + directory + file)
            	mrfs.append(directory + file)    
        if(os.path.isdir(directory + '\\' + file)):
            # print('Directory: ' + file)
            dir_search(directory + '\\' + file)

directory = 'C:\\Users\\nk91009578.NCOC\\Desktop\\MRFs\\2013\\'
dir_search(directory)

for mrf in mrfs:
	fl_label = [0,0]
	tl_label = [0,0]
	mrf_label = [0,0]
	date_label = [0,0]
	wb = load_workbook(filename = mrf)
	sheet = wb.active

	get_metadata_labels(sheet, fl_label, tl_label, mrf_label, date_label)

	fl = get_metadata_value(sheet, fl_label, tl_label[0])
	tl = get_metadata_value(sheet, tl_label, sheet.max_column + 1)
	mrf_n = get_metadata_value(sheet, mrf_label, date_label[0])	
	mrf_date = get_metadata_value(sheet, date_label, sheet.max_column + 1)
	if(type(mrf_date) is datetime.date or isinstance(mrf_date, datetime)):
		mrf_date = mrf_date.strftime('%d/%m/%Y')
	mrf_date = mrf_date.replace('.', '/')
	mrf_date = mrf_date.replace('-', '/')
	print(fl, tl, mrf_n, mrf_date)
