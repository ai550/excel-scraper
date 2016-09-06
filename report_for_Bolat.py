from openpyxl import load_workbook
from os.path import exists
import re

wb = load_workbook(filename = 'C:\\Users\\nk91009578.NCOC\\Desktop\\report_Bolat.xlsx')
root_dir = 'C:\\Users\\nk91009578.NCOC\\Desktop\\MRF'
sheet = wb.active
result_sheet = wb.get_sheet_by_name("Result")
row_index = 1
mrfs = []

for row in range(sheet.max_row):
    cell_name = "{}{}".format("A", row + 1)
    m = re.search('(?<=#)\w+', sheet[cell_name].value)
    a = m.group(0)


    print("MRF #" + a)
    if(exists(root_dir + '\\NCOC MRF ' + a + '.xlsx')):
        temp_wb = load_workbook(root_dir + '\\NCOC MRF ' + a + '.xlsx')
        temp_sheet = temp_wb.active
        for i in range(8, 31):
            d_cell_name = "{}{}".format("D", i)
            e_cell_name = "{}{}".format("E", i)
            c_n_1 = "{}{}".format("A", row_index)
            c_n_2 = "{}{}".format("B", row_index)
            c_n_3 = "{}{}".format("C", row_index)
            result_sheet[c_n_1] = temp_sheet[d_cell_name].value
            result_sheet[c_n_2] = "NCOC MRF#" + a
            row_index += 1
        wb.save('C:\\Users\\nk91009578.NCOC\\Desktop\\report_Bolat.xlsx')

    
