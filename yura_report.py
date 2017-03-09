from openpyxl import *
from time import process_time

report = load_workbook(filename="C:\\Users\\nk91009578.NCOC\\Desktop\\report_grande.xlsx")
objects = load_workbook(filename="C:\\Users\\nk91009578.NCOC\\Downloads\\object471.xlsx")

report_sheet = report.active
objects_sheet = objects.active

alist = []
uq_locations_list = []

monitor = "Monitor"
desktop = "Desktop"
workstation = "Workstation"
laptop = "Laptop"


def populateLists(alist, loc_list, report_sheet, objects_sheet):
    for cell in report_sheet.rows:
        alist.append(str(cell[0].value))
    alist.sort() # sort after population
    
    for cell in objects_sheet.rows:
        location = (str(cell[8].value)).strip() + ' ' + (str(cell[7].value)).strip()
        if location not in loc_list:
            loc_list.append(location.strip())
    loc_list.sort() # sort after population // possible bottleneck, optimize in future


"""
Binary Search algorithm
"""
def binarySearch(alist, item):
    first = 0 # inception point, here we define our first index of the alist to be 0
    last = len(alist)-1 # last index of given alist
    found = False # flag
    result = (False, 0) # result
    while first<=last and not found: # while first index is le last index and flag is set to False
        midpoint = (first + last)//2 # midpoint of given alist
        if alist[midpoint] == item: # if element with index midpoint in alist is our item
            found = True # then set flag to True, this ending the search
            result = (True, midpoint + 1) # give us the result
        else:
            if item < alist[midpoint]: # else, if our item is less than item at index midpoint in alist
                last = midpoint-1 # then make last equal to 
            else:
                first = midpoint+1
    return result


def mainProcess(alist, loc_list, sheet, r_sheet):

    cell_range = sheet['B2':'K' + str(sheet.max_row)]
    
    for cell in cell_range:
        name = str(cell[9].value)

        # check if name is present in report
        person_found = binarySearch(alist, name)
        location_found = binarySearch(loc_list, 
        
        """
        if result[0]:
            row = result[1]
            r_pc_hostname = r_sheet.cell(row=row, column=3) # PC Hostname
            r_pc_serial = r_sheet.cell(row=row, column=4)   # PC Serial
            r_pc_model = r_sheet.cell(row=row, column=5)    # PC Model
            r_scr1_model = r_sheet.cell(row=row, column=6)  # Screen 1 Model
            r_scr1_serial = r_sheet.cell(row=row, column=7) # Screen 1 Serial Number
            r_scr2_model = r_sheet.cell(row=row, column=8)  # Screen 2 Model
            r_scr2_serial = r_sheet.cell(row=row, column=9) # Screen 2 Serial Number
            r_comments = r_sheet.cell(row=row, column=10)   # Comments, additional information, etc...
            obj_type = str(cell[0].value)
            model = str(cell[2].value)
            serial_number = str(cell[3].value)
            if obj_type != monitor:
                hostname = str(cell[4].value)
                if r_pc_serial.value is None or r_pc_serial.value == '':
                    r_pc_serial.value = serial_number
                    r_pc_model.value = model
                    r_pc_hostname.value = hostname
                else:
                    if r_comments.value == None:
                        r_comments.value = 'Please be informed that user has additionally assigned items'
                    r_comments.value = '' + str(r_comments.value) + '\n{} SN: {} Model: {} Hostname: {}'.format(obj_type, serial_number, model, hostname)
            else:
                if r_scr1_serial.value is None or r_scr1_serial.value == '':
                    r_scr1_serial.value = serial_number
                    r_scr1_model.value = model
                elif r_scr2_serial.value is None or r_scr2_serial.value == '':
                    r_scr2_serial.value = serial_number
                    r_scr2_model.value = model
                else:
                    if r_comments.value == None:
                        r_comments.value = 'Please be informed that user has additionally assigned items'
                    r_comments.value = '' + str(r_comments.value) + '\n{} SN: {} Model: {}'.format(obj_type, serial_number, model)
        elif :
        """

populateLists(alist, uq_locations_list, report_sheet, objects_sheet)

print(uq_locations_list[0:10])

# Running
# mainProcess(alist, objects_sheet, report_sheet)

# report.save("C:\\Users\\nk91009578.NCOC\\Desktop\\grand_report.xlsx")


                
            

        
        
        
        
