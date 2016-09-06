from metadata import columns
from openpyxl import load_workbook

wb = load_workbook(filename='C:\\Users\\nk91009578.NCOC\\Desktop\\MRFs\\2013\\NCPOC MRF 0245.xlsx')
sheet = wb.active

field_dict = {'sn':'serial number', 'pn':'part number', 'ds':'description', 'q':'quantity', 'st':'status'}

for row in range(1, sheet.max_row):
	for col in range(1, sheet.max_column):
		if(col in columns):
			c_value = str(sheet[columns[col] + str(row)].value).strip().lower()
			if(c_value in field_dict.values()):
				print('found {} at {}'.format(c_value, columns[col] + str(row)))

# just cloned to my mbp
# gonna hack away
# another try