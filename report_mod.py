from openpyxl import load_workbook

wb1 = load_workbook(filename="C:\\Users\\nk91009578.NCOC\\Downloads\\CSD-ICTIM-ALL.xlsx")
sheet = wb1["Sheet1"]

"""
wb = load_workbook(filename="C:\\Users\\nk91009578.NCOC\\Documents\\report.xlsx")
sheet = wb.active

for row in range(2, sheet.max_row + 1):
    surname = "" if sheet["B" + str(row)].value == None else sheet["B" + str(row)].value
    name = "" if sheet["A" + str(row)].value == None else sheet["A" + str(row)].value
    sheet["C" + str(row)].value = surname + ", " + name

wb.save("C:\\Users\\nk91009578.NCOC\\Documents\\report.xlsx")
"""

for row in range(3, sheet.max_row+1):
    cell = sheet.cell(column=1, row=row)
    fixed_name = ""
    for i in list(str(cell.value)):
        if i != "[" and i != "(":
            fixed_name += i
        else:
            break
    new_cell = sheet.cell(column=2, row=row)
    new_cell.value = fixed_name.strip(" ")

"""
for row in range(2, sheet1.max_row + 1):
    desktop_counter = 0
    laptop_counter = 0
    workstation_counter = 0
    desktop_counter += 0 if sheet1["B" + str(row)].value == None else sheet1["B" + str(row)].value 
    laptop_counter += 0 if sheet1["C" + str(row)].value == None else sheet1["C" + str(row)].value
    workstation_counter += 0 if sheet1["D" + str(row)].value == None else sheet1["D" + str(row)].value

    sheet1["H" + str(row)].value = "User has {} desktop(s), {} laptop(s), {} workstation(s)".format(desktop_counter, laptop_counter, workstation_counter)

    dominant_obj_type = "Desktop"

    double_set_flag = "No"

    if desktop_counter > 1:
        double_set_flag = "Yes"
    if laptop_counter > desktop_counter:
        dominant_obj_type = "Laptop"
        if laptop_counter > 1:
            double_set_flag = "Yes"
    if workstation_counter > laptop_counter:
        dominant_obj_type = "Workstation"
        if workstation_counter > 1:
            double_set_flag = "Yes"
    
    sheet1["G" + str(row)].value = double_set_flag
    sheet1["F" + str(row)].value = dominant_obj_type
    
"""

wb1.save("C:\\Users\\nk91009578.NCOC\\Downloads\\CSD-ICTIM-ALL.xlsx")

print("DONE!")
